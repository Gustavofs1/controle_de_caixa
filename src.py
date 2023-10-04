from openpyxl import load_workbook
from  datetime import date, datetime
import time
import datetime
import os

book = load_workbook('Caixa.xlsx', data_only=True)
produtos = book['PRODUTOS']
entradas = book['ENTRADA']
saidas = book['SAIDA']
pedido_novo = True
hoje = str(date.today())
hoje = hoje + '.xlsx'

def adiciona_produto(pedido_novo):
    os.system('cls')
    print('---------AREA DE VENDAS------------')
    if pedido_novo == True:
        caixa_inicial = int(input('Valor do caixa: '))
        produtos['B3'] = caixa_inicial
    while True:
        sair = input('Adicionar um produto? (S/N): ')
        if sair.upper() == 'N':
            break
        if sair not in ['S', 'N', 's', 'N']:
            break
        else:
            horario = datetime.datetime.now()       #pega a hora atual
            horario = horario.strftime('%H:%M')
            produto = input('Nome do produto: ')
            try:
                quantidade = int(input('Quantidade: '))
                valor = float(input('Valor da venda: '))
            except:
                print('VOCÊ PRECISA DIGITAR UM VALOR NUMERICO!!')
                print('Retornando ao menu principal')
                time.sleep(2)
                os.system('python main.py')
                pass
            produtos.append([horario, produto, quantidade, valor])
            book.save('Caixa.xlsx')
            
        if sair == 'n' or sair == 'N':
            break
        else:
            pass
            


def fechar_caixa():
    os.system('cls')
    print('-----------AREA DE FECHAMENTO DO CAIXA-----------')
    print('')
    print('SEGUE INFORMAÇÕES DO CAIXA')
    print()
    print(f'caixa inicial foi: {produtos.cell(row=3, column=2).value}')
    print(f'Valor do caixa final: {produtos.cell(row=4, column=2).value}')
    print(f'Valor Total das Vendas: {produtos.cell(row=2, column=2).value}')
    print(f'Quantidade Total dos produtos vendidos: {produtos.cell(row=1, column=2).value}')
    print()
    print(f'Valor total das Saidas: {saidas.cell(row=2, column=4).value}')
    print(f'Valor total das Entradas: {entradas.cell(row=2, column=4).value}')
    print()
    limpar = input('Deseja encerrar o dia? (S/N): ')
    if limpar.upper() == 'S':
        while True:
            senha = input('Digite a sua senha para excluir, caso contrario digite SAIR para retornar: ')
            if senha.upper() == 'CENTERCELL':
                print('Seu caixa foi encerrado com sucesso!!! foi gerado um relatorio do dia na pasta, ao sair do sistema confira!')
                print('retornando ao menu principal...')
                time.sleep(5)
                book_backup = load_workbook('Caixa_backup.xlsx')
                book_backup.save(hoje)
                book_backup.save('Caixa.xlsx')
                break
            elif senha.upper() == 'SAIR':
                break
            else:
                print('Senha incorreta, tente novamente')
                print()
            
        
def menu():
    os.system('cls')
    print('''
            ------------------------------------------------------
                        CENTERCELL SYSTEM V1.0
            ------------------------------------------------------
          
            ********MENU PRINCIPAL********

            DIGITE UMA OPÇÃO
          
            1- Novo caixa:
            2- Continuar as vendas
            3- Fechar o dia
            4- Entrada de caixa
            5- Saida de caixa''')
    

def tela_entrada():
    os.system('cls')
    print('-----------AREA DAS ENTRADAS-----------')
    sair = input('Adicionar uma nova entrada? (S/N): ')
    if sair.upper() == 'N':
        pass
    if sair not in ['S', 'N', 's', 'N']:
        pass
    else:
        try:
            valor_entrada = float(input('Valor da entrada: '))
        except:
            print('VOCÊ PRECISA DIGITAR UM VALOR NUMERICO!!')
            print('Retornando ao menu principal...')
            time.sleep(2)    
        descricao_entrada = input('Descrição: ')
        tempo = datetime.datetime.now()       #pega a hora atual
        tempo = tempo.strftime('%H:%M')
        entradas.append([tempo, valor_entrada, descricao_entrada])
        book.save('Caixa.xlsx')
    
def tela_saida():
    os.system('cls')
    print('-----------AREA DAS SAIDAS-----------')
    sair = input('Adicionar uma nova saída? (S/N): ')
    if sair.upper() == 'N':
        pass
    if sair not in ['S', 'N', 's', 'N']:
        pass
    else:
        try:
            valor_saida = float(input('Valor da saida: '))
        except:
            print('VOCÊ PRECISA DIGITAR UM VALOR NUMERICO!!')
            print('Retornando ao menu principal...')
            time.sleep(2)
        descricao_saida = input('Descrição: ')
        tempo = datetime.datetime.now()       #pega a hora atual
        tempo = tempo.strftime('%H:%M')
        saidas.append([tempo, valor_saida, descricao_saida])
        book.save('Caixa.xlsx')